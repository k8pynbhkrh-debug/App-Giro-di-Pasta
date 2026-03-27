from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "tmp" / "generated_docs"
OUT.mkdir(parents=True, exist_ok=True)
TODAY = date(2026, 3, 27)

APP_JS = (ROOT / "app.js").read_text(encoding="utf-8")


def extract_balanced_block(text: str, marker: str) -> str:
    marker_index = text.find(marker)
    if marker_index < 0:
        raise RuntimeError(f"Marker not found: {marker}")

    start = marker_index + len(marker)
    while start < len(text) and text[start].isspace():
        start += 1

    open_char = text[start]
    close_char = {"[": "]", "{": "}"}[open_char]
    depth = 0
    in_string = False
    escape = False

    for index in range(start, len(text)):
        char = text[index]
        if in_string:
            if escape:
                escape = False
            elif char == "\\":
                escape = True
            elif char == '"':
                in_string = False
            continue

        if char == '"':
            in_string = True
        elif char == open_char:
            depth += 1
        elif char == close_char:
            depth -= 1
            if depth == 0:
                return text[start:index + 1]

    raise RuntimeError(f"Could not extract balanced block for {marker}")


def normalize_for_id(text: str) -> str:
    normalized = (
        text.lower()
        .replace("ä", "a")
        .replace("ö", "o")
        .replace("ü", "u")
        .replace("ß", "ss")
        .replace("’", " ")
        .replace("'", " ")
    )
    return (
        normalized.encode("ascii", "ignore")
        .decode("ascii")
        .replace("&", " und ")
        .replace("/", " ")
        .replace(",", " ")
        .replace("-", " ")
        .replace(".", " ")
        .replace("(", " ")
        .replace(")", " ")
        .strip()
    )


def normalize_key(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", normalize_for_id(text)).strip("_")


def sort_key(text: str) -> str:
    return normalize_for_id(text)


def parse_recipes_data() -> list[dict]:
    return json.loads(extract_balanced_block(APP_JS, "const recipesData = "))


def parse_recipe_guides() -> dict[str, dict]:
    return json.loads(extract_balanced_block(APP_JS, "const recipeGuidesById = "))


def parse_pdf_recipe_names() -> list[str]:
    match = re.search(r"const pdfRecipeNames = \[(.*?)\];", APP_JS, re.S)
    if not match:
        raise RuntimeError("Could not find pdfRecipeNames in app.js")
    return re.findall(r"'([^']+)'", match.group(1))


def count_recipe_images(folder: str, include_joker: bool = False) -> int:
    files = [path for path in (ROOT / folder).glob("*.png")]
    if not include_joker:
        files = [path for path in files if path.name != "joker.png"]
    return len(files)


def footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#6f6355"))
    canvas.drawRightString(doc.pagesize[0] - 18 * mm, 10 * mm, f"Stand {TODAY.isoformat()} - Seite {doc.page}")
    canvas.restoreState()


styles = getSampleStyleSheet()
TITLE = ParagraphStyle(
    "TitleCustom",
    parent=styles["Title"],
    fontName="Helvetica-Bold",
    fontSize=18,
    leading=22,
    textColor=colors.HexColor("#231b12"),
    alignment=TA_CENTER,
    spaceAfter=10,
)
SUBTITLE = ParagraphStyle(
    "SubtitleCustom",
    parent=styles["Heading2"],
    fontName="Helvetica-Bold",
    fontSize=13,
    leading=16,
    textColor=colors.HexColor("#6f3000"),
    spaceBefore=6,
    spaceAfter=6,
)
BODY = ParagraphStyle(
    "BodyCustom",
    parent=styles["BodyText"],
    fontName="Helvetica",
    fontSize=10,
    leading=14,
    textColor=colors.HexColor("#2c241d"),
    spaceAfter=5,
)
SMALL = ParagraphStyle(
    "SmallCustom",
    parent=BODY,
    fontSize=9,
    leading=12,
    textColor=colors.HexColor("#5f564d"),
)


def p(text: str, style=BODY):
    return Paragraph(text.replace("&", "&amp;"), style)


def bullet_lines(items: list[str], style=BODY):
    return [p(f"- {item}", style) for item in items]


def build_pdf(path: Path, story: list):
    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=16 * mm,
        title=path.stem,
        author="Codex for Giro: Pasta Night",
    )
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    reader = PdfReader(str(path))
    if len(reader.pages) == 0:
        raise RuntimeError(f"{path.name} has no pages")


GROUPED_RECIPES = [
    (
        "Klassiker und rote Saucen",
        [
            "Aglio e olio e peperoncino",
            "Pomodoro e basilico",
            "Arrabbiata",
            "Marinara",
            "Ricotta e pomodoro al peperoncino",
            "Cacio e pepe",
            "Carbonara",
            "Gricia",
            "Amatriciana",
            "Puttanesca",
            "Pasta all’Assassina",
        ],
    ),
    (
        "Sardellen, Kaese und cremige Saucen",
        [
            "Aglio, olio e acciughe",
            "Sardellen e limone",
            "Sardellen e pangrattato",
            "Burro e salvia",
            "Parmigiano e burro",
            "Funghi e burro",
            "Ricotta al limone",
            "Ricotta e gorgonzola",
            "Gorgonzola e noci",
            "Taleggio e pepe",
            "Stracchino e noci",
        ],
    ),
    (
        "Fleisch, Gefluegel und Wurst",
        [
            "Salsiccia e finocchio",
            "Speck e cipolla",
            "Pollo e panna",
            "Pollo e funghi",
        ],
    ),
    (
        "Fisch und Meeresfruechte",
        [
            "Tonno e pomodoro",
            "Tonno e capperi",
            "Salmone e panna",
            "Salmone e limone",
            "Gamberi e aglio",
            "Gamberi e zucchine",
            "Vongole in bianco",
            "Vongole e pomodoro",
            "Frutti di mare",
            "Bottarga e limone",
        ],
    ),
    (
        "Gemuese",
        [
            "Melanzane e pomodoro",
            "Zucchine e menta",
            "Spinaci e ricotta",
        ],
    ),
]

REMOVED_RECIPE_LINES = [
    "Pesto genovese",
    "Pesto rosso",
    "Ragu alla Bolognese",
    "Ragu bianco",
    "Quattro formaggi",
    "Crema di zucca",
    "Dessert-Rezepte aus dem alten V1-Katalog",
]


def validate_grouped_recipe_list(recipe_names: list[str]):
    flat = [name for _, names in GROUPED_RECIPES for name in names]
    if sorted(flat) != sorted(recipe_names):
        missing = sorted(set(recipe_names) - set(flat))
        extra = sorted(set(flat) - set(recipe_names))
        raise RuntimeError(f"Grouped recipe list mismatch. Missing={missing}, Extra={extra}")


def difficulty_dots(level: str) -> str:
    mapping = {
        "leicht": "●○○",
        "mittel": "●●○",
        "schwer": "●●●",
    }
    return mapping.get(level, "●○○")


def get_image_stats() -> tuple[int, int, int]:
    original_recipe_images = count_recipe_images("assets/recipes", include_joker=False)
    optimized_recipe_images = count_recipe_images("assets/recipes_optimized", include_joker=False)
    joker_images = 1 if (ROOT / "assets" / "recipes" / "joker.png").exists() else 0
    return original_recipe_images, optimized_recipe_images, joker_images


def build_architecture_pdf(recipe_names: list[str]):
    path = OUT / "App_Architektur_v1.4.pdf"
    original_images, optimized_images, joker_images = get_image_stats()

    story = [
        p("GIRO: PASTA NIGHT - App Architektur v1.4", TITLE),
        p("Diese Version aktualisiert v1.3 auf den aktuellen Stand der App inkl. Vorbereitungsschritt, lokalem QR, Web Share und aktuellem Rezeptkatalog.", SMALL),
        Spacer(1, 6),
        p("Technische Basis", SUBTITLE),
        *bullet_lines([
            "Single-page Web-App mit HTML, CSS und Vanilla JavaScript.",
            "Zentrale Laufzeitdateien bleiben index.html, app.js, manifest.json und service-worker.js.",
            "Gespeicherte Spiele bleiben in localStorage; es gibt weiterhin keinen Pflicht-Backend-Flow fuer den normalen Spielbetrieb.",
            "Der Spectator arbeitet mit einem kompakten Watch-Snapshot in der URL; QR-Codes werden lokal im Browser erzeugt und haben nur noch einen externen Fallback.",
        ]),
        p("Zentraler Spielzustand", SUBTITLE),
        *bullet_lines([
            "Jedes Spiel besitzt u. a. id, title, phase, gameMode, settings, rounds, players, gameIndex, activePlayerTurnIndex, scores, shoppingList und finished.",
            "gameMode ist ein First-class-State mit den Werten 'guessing' oder 'open' und steuert Sichtbarkeit, Score-Verhalten, Spectator-Darstellung und Summary-Flow zentral.",
            "Alte Savegames werden beim Laden normalisiert; entfernte Rezepte werden bereinigt und die Einkaufsliste neu berechnet.",
        ]),
        p("Hauptfluss der App", SUBTITLE),
        *bullet_lines([
            "Landing / Spielverwaltung",
            "Konfiguration mit Spielerzahl, Rezepte pro Spiel, Spielmodus und Lebensmittel-Filtern",
            "Summary / Planung mit Einkaufsliste und Modus-spezifischem Rezeptzugriff",
            "Vorbereitung als eigener Schritt vor dem Spieler-Setup",
            "Spieler-Setup",
            "Game Screen mit Reveal im Modus Mit Raten bzw. direkter Rezeptansicht im offenen Modus",
            "Spielrueckblick mit nur den tatsaechlich gespielten Rezepten ohne Joker",
        ]),
        p("Zwei sauber getrennte Spielmodi", SUBTITLE),
        *bullet_lines([
            "Mit Raten (empfohlen): Rezeptplan bleibt in der Summary verborgen, Reveal-Screen schuetzt vor Spoilern, Punkte bleiben waehrend der Runde aktiv.",
            "Ohne Raten: Rezeptplan ist in der Summary sichtbar und editierbar, Score-Logik bleibt komplett aus, der Spectator zeigt nur die geplanten und aktuellen Rezepte.",
            "Die Filterlogik fuer Fleisch, Fisch, tierische Produkte und Scharf gilt in beiden Modi identisch vor der Rezeptauswahl.",
        ]),
        PageBreak(),
        p("Rezept- und Planungslogik", SUBTITLE),
        *bullet_lines([
            "Aktiver Rezeptkatalog: 39 Saucen aus dem aktuellen Stand von Pasta Saucen v2.3; Joker-Runden werden separat erzeugt.",
            "Pro Spiel werden ca. 20 Prozent Joker-Runden eingestreut, der Rest kommt aus dem gefilterten Rezeptpool.",
            "Im offenen Modus koennen vorgeschlagene Rezepte entfernt und beliebig viele Zusatzrezepte bestaetigt hinzugefuegt werden - entweder passend zu den Filtern oder bewusst ausserhalb der Filter.",
            "Die Einkaufsliste aggregiert alle Nicht-Joker-Runden auf Basis von ca. 25 g Pasta pro Person pro Runde und ergaenzt eine pauschale Empfehlung fuer Kuechenpapier / Kuechenrolle.",
            "Nach der Einkaufsliste folgt eine eigene Vorbereitungsseite mit deduplizierten Zutaten fuer 'Vorbereiten' und 'Bereitstellen'.",
        ]),
        p("Score-, Export- und Rueckblicklogik", SUBTITLE),
        *bullet_lines([
            "Punkte gelten nur im Modus Mit Raten: erster richtiger Tipp +2, jeder falsche Tipp -1, Koch +3 nur wenn niemand richtig liegt.",
            "Waehrend des Kochens bleiben Score-Anzeige und Score-Bedienung sichtbar; in Ohne Raten wird Score nirgends berechnet oder angezeigt.",
            "Einkaufsliste kann per Web Share API, Clipboard, Bring-kompatibler Copy-Liste, WhatsApp und Druck / PDF weitergegeben werden.",
            "Der Spielrueckblick listet nur die wirklich gespielten Rezepte ohne Joker in tatsaechlicher Reihenfolge; von dort aus gibt es auch den WhatsApp-Export inklusive Kochanleitungen.",
        ]),
        p("Spectator, Bilder und PWA", SUBTITLE),
        *bullet_lines([
            "Spectator-Ansicht ist read-only und wird ueber QR-Link mit Watch-Snapshot geoeffnet.",
            "Mit Raten zeigt der Spectator Scores plus den kompletten verfuegbaren Rezeptkatalog; Ohne Raten zeigt er nur die tatsaechlich gespielten Rezepte und markiert das aktuelle Rezept.",
            f"Es liegen aktuell {original_images} Rezeptbilder plus {joker_images} Joker-Bild im Projekt; die App lädt bevorzugt {optimized_images} optimierte Web-Bilder aus assets/recipes_optimized mit Fallback auf die Originale.",
            "Der Service Worker cached App Shell, Manifest, Icons, lokalen QR-Generator und die Bild-Assets fuer schnellere Wiederaufrufe.",
        ]),
        p("Bewusst nicht im Scope", SUBTITLE),
        *bullet_lines([
            "Kein Timer, keine Countdown-Logik, keine Timer-Sounds und keine timergetriebene Haptik.",
            "Kein Framework, kein serverseitiger Pflicht-Backend-Flow und weiterhin keine IndexedDB-Migration.",
            "Bildgenerierung bleibt extern; die App verwaltet nur Dateinamen, Slots und optimierte Web-Kopien.",
        ]),
        p(f"Aktive Rezeptanzahl ohne Joker: {len(recipe_names)}", SMALL),
    ]
    build_pdf(path, story)


def build_concept_pdf():
    path = OUT / "App_Konzept_v1.4.pdf"
    original_images, optimized_images, joker_images = get_image_stats()

    story = [
        p("GIRO: PASTA NIGHT - App Konzept v1.4", TITLE),
        p("Aktualisierte Produktbeschreibung fuer den aktuellen Stand ohne Timer, mit Vorbereitungsschritt und klar getrennten Spielmodi.", SMALL),
        Spacer(1, 6),
        p("Grundidee", SUBTITLE),
        *bullet_lines([
            "Giro: Pasta Night ist ein gemeinsames Kochspiel fuer den Tisch: Das Smartphone ersetzt Rezeptkarten, Einkaufsliste, Vorbereitungszettel und Teile der Spielleitung.",
            "Der Koch liest nur die Schritte. Geraten wird muendlich zwischen den Mitspielern, nicht in der App.",
            "Die App bleibt auch mit fettigen oder nassen Haenden bedienbar: grosse Buttons, klare Ueberschriften, wenig Ablenkung und konsistente Platzierung der Weiter-Aktionen.",
        ]),
        p("Konfiguration", SUBTITLE),
        *bullet_lines([
            "Spieler- bzw. Gruppenanzahl (bis 6 spielende Parteien).",
            "Rezepte pro Spiel mit dem Hinweis auf ca. 25 g Pasta pro Person pro Runde.",
            "Spielmodus-Auswahl: Mit Raten (empfohlen) oder Ohne Raten.",
            "Filter fuer Fleisch, Fisch, tierische Produkte und Scharf.",
        ]),
        p("Spielmodus 1 - Mit Raten", SUBTITLE),
        *bullet_lines([
            "Die App waehlt passende Rezepte intern aus, ohne sie in Summary oder Vorbereitung preiszugeben.",
            "Vor jeder Runde schuetzt ein neutraler Handover-Screen den geheimen Rezeptnamen.",
            "Punkte bleiben nur waehrend der Runden aktiv; der Abschluss zeigt trotzdem nur die wirklich gespielten Rezepte.",
            "Der Spectator hilft hier beim Raten mit Scores und dem gesamten verfuegbaren Rezeptkatalog.",
        ]),
        p("Spielmodus 2 - Ohne Raten", SUBTITLE),
        *bullet_lines([
            "Die App zeigt den vorgeschlagenen Rezeptplan offen in der Summary.",
            "Der Organisator kann Rezepte loeschen und beliebig viele Zusatzrezepte bestaetigt hinzufuegen - passend zu den Filtern oder bewusst ausserhalb davon.",
            "Es gibt keine Punkte, keine Rangliste und keinen Wettbewerbsvorteil ueber Guessing.",
            "Der Spectator dient hier nur als offene Rezeptuebersicht fuer die tatsaechlich gespielten Runden.",
        ]),
        PageBreak(),
        p("Summary, Vorbereitung und Export", SUBTITLE),
        *bullet_lines([
            "Mit Raten zeigt die Summary nur Einkaufsliste und neutrale Auswahl-Info; Ohne Raten zeigt den editierbaren Rezeptplan.",
            "Nach der Einkaufsliste folgt eine eigene Vorbereitungsseite mit deduplizierten Zutatenlisten fuer 'Vorbereiten' und 'Bereitstellen'.",
            "Exportwege fuer den Einkauf: Web Share / Apple Erinnerungen, Clipboard, Bring-kompatible Liste, WhatsApp und Druck / PDF.",
            "Der Rueckblick exportiert die gespielten Rezepte inklusive Kochanleitungen per WhatsApp.",
        ]),
        p("Gameplay UX", SUBTITLE),
        *bullet_lines([
            "Reveal-Screen fuer geheimen Handy-Handover im Wettbewerbsmodus.",
            "Kochscreen mit Gerichtstitel, Bild, Schrittliste, Score-Anzeige, Skip-Button sowie kompakter Schwierigkeit und Tipp.",
            "Haptisches Feedback auf Neue Spiele, Rezept anzeigen, Rezept skippen und Bestaetigen.",
            "Spiel neu starten setzt das aktuelle Spiel wieder auf Runde 1 zurueck, ohne neue Rezepte oder neue Einkaufsliste zu erzeugen.",
        ]),
        p("Bilder, Zuschauer und PWA", SUBTITLE),
        *bullet_lines([
            f"Aktuell sind {original_images} Rezeptbilder plus {joker_images} Joker-Bild eingebunden; fuer die Web-Auslieferung liegen {optimized_images} optimierte Kopien vor.",
            "QR-Codes werden lokal in der App erzeugt; nur falls das fehlschlaegt, bleibt ein externer Fallback moeglich.",
            "Die App bleibt offline-grundfaehig ueber Service Worker, Manifest und gecachte Kern-Assets.",
        ]),
        p("Produktgrenzen im aktuellen Stand", SUBTITLE),
        *bullet_lines([
            "Kein Timer, keine Audio-Signale und keine Timer-Farbwechsel.",
            "Keine IndexedDB-Migration und kein Pflicht-Backend fuer normale Spiele.",
            "Bildgenerierung bleibt extern; die App verwaltet nur die Einbindung.",
        ]),
    ]
    build_pdf(path, story)


def build_architecture_card_pdf():
    path = OUT / "Architekturkarte v1.3.pdf"
    story = [
        p("GIRO: PASTA NIGHT - Architekturkarte v1.3", TITLE),
        p("Kompakte Uebersicht des aktuellen App-Flows", SMALL),
        Spacer(1, 6),
    ]

    data = [
        ["1. Landing", "Spiel anlegen, laden, erneut planen, von vorne beginnen oder loeschen"],
        ["2. Konfiguration", "Spielerzahl, Rezepte pro Spiel, Spielmodus und Filter"],
        ["3. Generierung", "Rezepte aus dem erlaubten Pool plus Joker-Runden erzeugen"],
        ["4a. Summary - Mit Raten", "Nur Einkaufsliste und neutrale Auswahl-Info, keine Rezeptnamen"],
        ["4b. Summary - Ohne Raten", "Rezeptplan sichtbar, Rezepte entfernen und Zusatzrezepte bestaetigt ergaenzen"],
        ["5. Vorbereitung", "Deduplizierte Zutaten fuer Vorbereiten und Bereitstellen"],
        ["6. Spieler-Setup", "Spielernamen erfassen, Reihenfolge mischen, Kochspiel starten"],
        ["7a. Runde - Mit Raten", "Nächster Koch -> Rezept anzeigen -> Score / Skip -> Runde beenden"],
        ["7b. Runde - Ohne Raten", "Direkt Rezeptansicht -> Skip optional -> Runde beenden"],
        ["8. Spielrueckblick", "Nur tatsaechlich gespielte Rezepte ohne Joker; optional WhatsApp-Export der Kochanleitungen"],
        ["Shared", "Autosave, lokale QR-Erzeugung, Spectator, Vorbereitung, PWA / Offline, Bildslot unter dem Titel"],
    ]
    table = Table(data, colWidths=[52 * mm, 112 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f2ebdf")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#d9c9b5")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e7dfd1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("LEADING", (0, 0), (-1, -1), 13),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    story.extend([
        Spacer(1, 12),
        p("Wichtige Architekturregeln", SUBTITLE),
        *bullet_lines([
            "Der Spielmodus trennt Sichtbarkeit, Score-Verhalten und Spectator-Darstellung eindeutig.",
            "Mit Raten leakt den konkreten Plan in Summary und Vorbereitung nicht an den Organisator.",
            "Ohne Raten rechnet keine Punkte im Hintergrund mit.",
            "Der Abschluss fokussiert nur noch den tatsaechlich gespielten Rezeptpfad statt eine separate Siegerseite zu bauen.",
        ]),
    ])
    build_pdf(path, story)


def build_rules_pdf():
    path = OUT / "Regelblatt_v1.4.pdf"
    story = [
        p("GIRO: PASTA NIGHT - Regelblatt v1.4", TITLE),
        p("Aktuelle Spielregeln ohne Timer und ohne Abschluss-Rangliste.", SMALL),
        Spacer(1, 6),
        p("Gemeinsame Grundregeln", SUBTITLE),
        *bullet_lines([
            "Maximal 6 spielende Parteien. Ab 6 Personen werden Teams gebildet; jedes Team zaehlt als eine Partei.",
            "Die Rezeptschritte sind nicht interaktiv. Der Koch liest und kocht, die anderen sprechen ihre Vermutungen nur muendlich aus.",
            "Joker-Runden koennen im Spiel auftauchen und stehen fuer eine freie Wahl statt eines festen Rezepts.",
            "Rezept skippen bleibt als sichtbare Aktion waehrend des Kochens erlaubt.",
            "Vor dem Spieler-Setup gibt es eine Vorbereitungsseite fuer alle deduplizierten Tisch- und Vorbereitungsaufgaben.",
        ]),
        p("Ablauf einer Runde - Mit Raten", SUBTITLE),
        *bullet_lines([
            "Das Smartphone wird an den naechsten Koch uebergeben.",
            "Der Reveal-Screen zeigt nur Nächster Koch und den Namen der aktiven Person.",
            "Erst nach Tipp auf Rezept anzeigen sieht der Koch das Rezept.",
            "Andere Spieler raten muendlich, waehrend der Koch arbeitet.",
            "Die Punkte werden direkt auf dem Kochscreen vergeben.",
        ]),
        p("Punktevergabe - nur Mit Raten", SUBTITLE),
        *bullet_lines([
            "Erster richtiger Tipp: +2 Punkte.",
            "Jeder falsche Tipp: -1 Punkt.",
            "Niemand errät das Rezept: Koch +3 Punkte.",
            "Mindestens ein richtiger Tipp: Koch 0 Punkte.",
        ]),
        PageBreak(),
        p("Ablauf einer Runde - Ohne Raten", SUBTITLE),
        *bullet_lines([
            "Die Rezepte sind bereits in der Planung sichtbar.",
            "Das Spiel laeuft offen, ohne Reveal-Pflicht und ohne Score-System.",
            "Zusatzrezepte duerfen im offenen Modus auch bewusst ausserhalb der urspruenglichen Filter hinzugefuegt werden.",
        ]),
        p("Spielende", SUBTITLE),
        *bullet_lines([
            "Der Spielrueckblick zeigt nur die tatsaechlich gespielten Rezepte ohne Joker in ihrer echten Reihenfolge.",
            "Von dort aus koennen die gespielten Rezepte inklusive Kochanleitungen per WhatsApp exportiert werden.",
            "Spiel neu starten setzt dasselbe Spiel wieder auf Runde 1 zurueck, ohne neue Rezepte zu generieren.",
        ]),
        p("Zuschauermodus", SUBTITLE),
        *bullet_lines([
            "Der Spectator wird per QR-Code geoeffnet und bleibt read-only.",
            "Mit Raten zeigt er Punkte sowie den gesamten verfuegbaren Rezeptkatalog als Hilfestellung fuer die Ratenden.",
            "Ohne Raten zeigt er nur die wirklich gespielten Rezepte und markiert das aktuelle Rezept.",
        ]),
        p("Wichtiger Hinweis", SUBTITLE),
        *bullet_lines([
            "Der fruehere Timer ist vollstaendig aus dem Produkt entfernt. Es gibt keine Countdown-Regel mehr.",
        ]),
    ]
    build_pdf(path, story)


def build_recipe_list_pdf(recipe_names: list[str]):
    path = OUT / "Rezeptliste v1.3.pdf"
    validate_grouped_recipe_list(recipe_names)
    story = [
        p("GIRO: PASTA NIGHT - Rezeptliste v1.3", TITLE),
        p("Aktiver Rezeptkatalog der App auf Basis von Pasta Saucen v2.3.", SMALL),
        Spacer(1, 6),
        p(f"Aktive Rezepte: {len(recipe_names)}. Joker-Runden werden separat erzeugt und sind kein eigenes Rezept.", BODY),
    ]
    for title, names in GROUPED_RECIPES:
        story.append(p(title, SUBTITLE))
        story.extend(bullet_lines(names))
    story.extend([
        Spacer(1, 8),
        p("Nicht mehr im aktiven Katalog: Pesto-Rezepte, Dessert-Rezepte, Ragus, Quattro formaggi und Crema di zucca.", SMALL),
    ])
    build_pdf(path, story)


def build_task_list_xlsx():
    path = OUT / "Aufgabenliste v1.4.xlsx"
    rows = [
        (1, "UX", "Privacy Screen beim Handy-Handover", "Rezept erst nach Uebergabe sichtbar machen", "Erledigt", "Reveal-Screen aktiv im Modus Mit Raten."),
        (2, "UX", "Cooking Screen bereinigen", "Titel, Schritte, Score, Skip, Schwierigkeit und Tipp sauber darstellen", "Erledigt", "Timer-UI entfernt, Kochscreen bleibt praktisch."),
        (3, "UX", "Spielmodus Mit Raten", "Verdeckte Rezeptplanung und Wettbewerb aktiv halten", "Erledigt", "Summary versteckt Rezeptnamen, Reveal und Score aktiv."),
        (4, "UX", "Spielmodus Ohne Raten", "Offene Planung ohne Wettbewerb", "Erledigt", "Rezeptplan sichtbar und editierbar, Score komplett aus."),
        (5, "UX", "Big Tap Targets", "Alle kritischen Buttons mobil gut treffbar machen", "Teilweise", "Hauptaktionen sind gross; einige Randaktionen koennen spaeter noch weiter vereinheitlicht werden."),
        (6, "UX", "Kompakte Score-Anzeige", "Punktestand sichtbar halten ohne den Kochscreen zu ueberladen", "Erledigt", "Aufklappbare Ranglistenansicht im Hauptspiel vorhanden."),
        (7, "Feedback", "Haptik fuer Schluesselaktionen", "Kurzes Vibrationsfeedback bei wichtigen Buttons", "Erledigt", "Neu, Reveal, Skip, Bestaetigen und Score nutzen navigator.vibrate mit Fallback."),
        (8, "Feedback", "Timer-Sound / Timer-Vibration / Timer-HUD", "Fruehere Countdown-Features", "Aus Scope", "Der Timer wurde bewusst aus dem Produkt entfernt."),
        (9, "Stabilitaet", "Autosave implementieren", "Spielstaende automatisch speichern", "Erledigt", "Speicherung erfolgt ueber upsertCurrentGame in localStorage."),
        (10, "Datenstruktur", "Rezept-Datenmodell strukturieren", "Zutaten, IDs, Guides und Bilder konsistenter zusammenfuehren", "Teilweise", "Runtime-Daten liegen noch in app.js und recipes.json verteilt."),
        (11, "Datenstruktur", "IndexedDB statt localStorage", "Robustere Speicherung", "Offen", "Persistenz laeuft weiterhin ausschliesslich ueber localStorage."),
        (12, "PWA", "Service Worker", "Offlinefaehigkeit der App", "Erledigt", "service-worker.js ist aktiv und versioniert."),
        (13, "PWA", "App Shell Caching", "Schneller App-Start und Fallbacks", "Erledigt", "HTML, JS, Manifest, Icons, lokaler QR-Generator und Bilder werden gecacht."),
        (14, "Feature", "QR Spectator Mode", "Read-only Zuschaueransicht ueber QR", "Erledigt", "Watch-Link, lokaler QR-Code und Snapshot-Fallback sind produktiv."),
        (15, "Feature", "Spectator mode-aware UI", "Anzeige an Mit Raten / Ohne Raten anpassen", "Erledigt", "Mit Raten zeigt Scores und alle Rezepte, Ohne Raten nur gespielte Rezepte."),
        (16, "Planung", "Einkaufsliste generieren", "Zutaten automatisch aggregieren", "Erledigt", "Nicht-Joker-Runden werden fuer die Einkaufsliste zusammengefasst."),
        (17, "Integration", "Bring-friendly Export", "Einkaufsliste besser fuer Bring nutzbar machen", "Erledigt", "Eigener Copy-for-Bring-Button liefert minimales Zeilenformat."),
        (18, "Integration", "Web Share API", "Liste nativ ueber das Betriebssystem teilen", "Erledigt", "Apple Erinnerungen / nativer Share-Sheet mit Clipboard-Fallback."),
        (19, "Integration", "Apple Erinnerungen Export", "iOS-kompatiblen Listenfluss anbieten", "Erledigt", "Share Sheet plus sauberer Fallback auf Clipboard-Text."),
        (20, "Gameplay", "Rezeptplanung offen editieren", "Rezepte entfernen und Zusatzrezepte bestaetigen", "Erledigt", "Nur im Modus Ohne Raten sichtbar und nutzbar."),
        (21, "Gameplay", "Zusatzpunkt bestes Gericht", "Optionaler Bonuspunkt pro Runde", "Offen", "Keine UI und keine State-Logik vorhanden."),
        (22, "Gameplay", "Rezepte nach Spiel bereitstellen", "Saubere Nach-Spiel-Uebersicht aller gekochten Gerichte", "Erledigt", "Spielrueckblick listet nur tatsaechlich gespielte Rezepte ohne Joker."),
        (23, "Content", "Rezeptkatalog V2.2 synchronisieren", "Aktive Rezepte auf aktuellen Dokumentenstand bringen", "Erledigt", "Pesto, Dessert-Rezepte und Crema di zucca sind entfernt; Schritte auf aktuellem Sprachstand."),
        (24, "Rezepte", "Puttanesca pruefen", "Fachliche Pruefung dokumentieren", "Erledigt", "Review dokumentiert; Rezepttext sprachlich eingegrenzt, ohne neue Kochlogik zu erfinden."),
        (25, "Rezepte", "Amatriciana Zwiebel klaeren", "Aromaten-Frage fachlich final entscheiden", "Teilweise", "Aktueller Stand ist dokumentiert, aber die kulinarische Endentscheidung bleibt bewusst offen."),
        (26, "Rezepte", "Pasta all'Assassina validieren", "Fachliche Beschreibung final absichern", "Teilweise", "Review vorhanden, Endentscheidung noch offen."),
        (27, "Content", "Rezeptbilder Hook und Vollbestand", "Bild-Slot unter dem Titel und externe Assets nutzen", "Erledigt", "39 Rezeptbilder plus Joker sind vorhanden; optimierte Web-Bilder werden bevorzugt geladen."),
        (28, "Produkt", "Asia Edition Konzept", "Alternative Rezeptbibliothek", "Verschoben", "Aktuell nicht im Fokus der App-Umsetzung."),
        (29, "Business", "Zutatenpartner / Rabatt / Commission", "Partner- und Erlosmodell bewerten", "Extern", "Kein Code-Thema, bleibt ausserhalb des Produktkerns."),
        (30, "Flow", "Vorbereitungsseite", "Deduplizierte Aufgaben zwischen Einkauf und Spielerauswahl", "Erledigt", "Eigener Schritt im Flow mit Vorbereiten/Bereitstellen."),
        (31, "Flow", "Rezeptexport nach Spiel", "Gespielte Rezepte inklusive Kochanleitungen teilen", "Erledigt", "WhatsApp-Export am Spielende vorhanden."),
        (32, "Flow", "Spiel neu starten", "Dasselbe Spiel wieder ab Runde 1 spielen", "Erledigt", "Setzt dasselbe Spiel auf Runde 1 zurueck."),
        (33, "Technik", "Lokale QR-Erzeugung", "QR nicht mehr primär extern laden", "Erledigt", "Lokaler QR-Generator ist eingebunden, externer Dienst nur noch Fallback."),
        (34, "Technik", "Bildoptimierung und Sync", "Web-Bilder kleiner ausliefern und Assets komfortabel nachziehen", "Erledigt", "Optimierte Rezeptbilder, Watcher und Fast-Deploy-Skripte vorhanden."),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Giro Pasta Night Tasks v1.4"
    ws.append(["Prioritaet", "Kategorie", "Aufgabe", "Beschreibung / Ziel", "Status", "Stand / Hinweis"])
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="F2EBDD")
    status_fills = {
        "Erledigt": PatternFill("solid", fgColor="DCEED7"),
        "Teilweise": PatternFill("solid", fgColor="FFF0C9"),
        "Offen": PatternFill("solid", fgColor="F8D6D2"),
        "Verschoben": PatternFill("solid", fgColor="E7E1F6"),
        "Extern": PatternFill("solid", fgColor="D9E7F7"),
        "Aus Scope": PatternFill("solid", fgColor="E6E6E6"),
    }
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status_cell = row[4]
        status_cell.fill = status_fills.get(status_cell.value, PatternFill())
        status_cell.font = Font(bold=True)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {1: 10, 2: 16, 3: 34, 4: 44, 5: 14, 6: 52}
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    wb.save(path)


def get_current_ingredients(recipes_data: list[dict]) -> list[str]:
    labels = {"Pasta (g)"}
    for recipe in recipes_data:
        labels.update(recipe["ingredients"].keys())
    return ["Pasta (g)"] + sorted((label for label in labels if label != "Pasta (g)"), key=sort_key)


def ingredient_base_label(label: str) -> str:
    base = re.sub(r"\s*\([^)]+\)\s*$", "", label).strip()
    return {
        "Zitrone (Saft)": "Zitrone",
    }.get(base, base)


def build_amount_model_xlsx(recipes_data: list[dict]):
    path = OUT / "Mengenmodell Portionsbasis v1.4.xlsx"
    ingredient_columns = get_current_ingredients(recipes_data)
    recipe_count = len(recipes_data)
    default_rounds = 8

    def recipe_row(recipe: dict, scale: float, pasta_amount: float) -> list:
        data = {"Pasta (g)": pasta_amount}
        for ingredient, amount in recipe["ingredients"].items():
            data[ingredient] = amount * scale
        return [recipe["name"]] + [data.get(column, 0) for column in ingredient_columns]

    recipe_rows_25 = [recipe_row(recipe, 1, 25) for recipe in recipes_data]
    recipe_rows_100 = [recipe_row(recipe, 4, 100) for recipe in recipes_data]

    totals = {}
    for column in ingredient_columns:
        if column == "Pasta (g)":
            totals[column] = 25
            continue
        totals[column] = sum(recipe["ingredients"].get(column, 0) for recipe in recipes_data) / recipe_count

    wb = Workbook()
    ws_assumptions = wb.active
    ws_assumptions.title = "Assumptions"
    assumption_rows = [
        ("Parameter", "Wert", "Erläuterung"),
        ("Anzahl Rezepte", recipe_count, "aktueller App-Stand / Pasta Saucen v2.3"),
        ("Pasta pro Rezept & Person (g)", 25, "freie Pastasorte, Menge zählt pro Runde"),
        ("Referenzbasis (g Pasta)", 100, "Alle Rezeptmengen beziehen sich zusätzlich auf 100 g Pasta"),
        ("Skalierungsfaktor 100g→25g", 0.25, "25/100"),
        ("Hinweis", "Bei 8 Rezepten und 4 Personen wären das 8×25g×4 = 800g Pasta.", "Die Pastasorte bleibt frei wählbar."),
        ("Rezepte gekocht (N)", default_rounds, "Für den Abend: Gesamt = Portion × N × Personen. N ist editierbar."),
        ("Interpretation Totals_1p", "1 Portion", f"Durchschnitt pro Rezeptkarte (uniform über {recipe_count} Rezepte) für 25 g Pasta."),
        ("Nicht quantifiziert", "Küchenpapier / Küchenrolle", "Empfehlung erscheint in der App-Einkaufsliste, ist aber kein Mengenmodell-Wert."),
        ("Version", "1.4", "Auf aktuellen V2.3-Rezeptstand und aktuelle Zutaten bereinigt am 2026-03-27"),
    ]
    for row in assumption_rows:
        ws_assumptions.append(row)

    def append_recipe_sheet(title: str, rows: list[list]):
        ws = wb.create_sheet(title)
        ws.append(["Rezept"] + ingredient_columns)
        for row in rows:
            ws.append(row)
        return ws

    ws_100 = append_recipe_sheet("Recipe_100g", recipe_rows_100)
    ws_25 = append_recipe_sheet("Recipe_25g_1p", recipe_rows_25)

    ws_totals = wb.create_sheet("Totals_1p")
    ws_totals.append(["Zutat", "Gesamtmenge für 1 Person", "Herleitung (auf Basis 100g→25g)"])
    for ingredient in ingredient_columns:
        value = totals[ingredient]
        if ingredient == "Pasta (g)":
            derivation = "25 (feste Menge pro Rezept und Person)"
        else:
            total_100 = value * recipe_count * 4
            derivation = f"0.25×({round(total_100, 4)}/{recipe_count})"
        ws_totals.append([ingredient, round(value, 4), derivation])

    ws_scale = wb.create_sheet("Scale_1-10p")
    ws_scale.append(["Personen"] + ingredient_columns)
    for people in range(1, 11):
        ws_scale.append([people] + [round(totals[column] * people, 4) for column in ingredient_columns])

    ws_scale_n = wb.create_sheet("Scale_1-10p_NRezepte")
    ws_scale_n.append(["Personen", "Rezepte (N)"] + ingredient_columns)
    for people in range(1, 11):
        ws_scale_n.append([people, default_rounds] + [round(totals[column] * people * default_rounds, 4) for column in ingredient_columns])

    header_fill = PatternFill("solid", fgColor="F2EBDD")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        widths = defaultdict(lambda: 14)
        widths[1] = 26
        if ws.title == "Assumptions":
            widths[1] = 24
            widths[2] = 34
            widths[3] = 48
        elif ws.title == "Totals_1p":
            widths[1] = 24
            widths[2] = 18
            widths[3] = 30
        elif ws.title == "Scale_1-10p_NRezepte":
            widths[1] = 10
            widths[2] = 12
        elif ws.title == "Scale_1-10p":
            widths[1] = 10

        for idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = widths[idx]

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(path)


def build_ingredient_list_pdf(recipes_data: list[dict]):
    path = OUT / "Zutatenliste v1.2.pdf"
    ingredient_set = {ingredient_base_label(label) for label in get_current_ingredients(recipes_data)}

    grouped = [
        ("Immer dabei", ["Pasta (freie Sorte)", "Küchenpapier / Küchenrolle (Empfehlung)"]),
        ("Basis und Vorrat", ["Olivenöl", "Butter", "Schwarzer Pfeffer", "Oregano", "Pangrattato", "Weißwein"]),
        ("Gemüse, Kräuter und Zitrus", ["Knoblauch", "Zwiebeln", "Tomaten", "Chili/Peperoncino", "Champignons", "Fenchel", "Zucchini", "Aubergine", "Spinat", "Basilikum", "Petersilie", "Minze", "Salbei", "Zitrone"]),
        ("Käse, Milchprodukte und Eier", ["Eier", "Milch", "Sahne", "Ricotta", "Parmigiano", "Pecorino", "Gorgonzola", "Taleggio", "Stracchino"]),
        ("Fisch und Meeresfrüchte", ["Sardellen", "Thunfisch", "Lachs", "Garnelen", "Vongole", "Meeresfrüchte-Mix", "Bottarga"]),
        ("Fleisch, Wurst und Geflügel", ["Guanciale/Pancetta", "Speck", "Salsiccia", "Hähnchen"]),
        ("Extras", ["Kapern", "Oliven", "Walnüsse"]),
    ]

    missing = sorted(
        ingredient for ingredient in ingredient_set
        if ingredient not in {item.replace(" (freie Sorte)", "").replace(" (Empfehlung)", "") for _, items in grouped for item in items}
    )
    if missing:
        raise RuntimeError(f"Zutatenliste enthält noch nicht zugeordnete Zutaten: {missing}")

    story = [
        p("GIRO: PASTA NIGHT - Zutatenliste v1.2", TITLE),
        p("Aktualisierte Zutatenuebersicht passend zum aktuellen Rezeptkatalog und zur generischen Pasta-Regel.", SMALL),
        Spacer(1, 6),
        p("Hinweis: Die App rechnet immer mit generischer Pasta. Welche Form gekauft wird, bleibt bewusst offen.", BODY),
    ]
    for title, items in grouped:
        story.append(p(title, SUBTITLE))
        story.extend(bullet_lines(items))
    build_pdf(path, story)


def build_pasta_sauces_text(recipe_names: list[str], recipe_guides: dict[str, dict]):
    path = OUT / "Pasta Saucen v2.3.txt"
    lines = [
        "Giro: Pasta Night – finale pfannentaugliche Rezeptliste v2.3",
        "",
        "Stand: 2026-03-26",
        "",
        "Entfernt / nicht mehr im aktiven Spielkatalog:",
    ]
    lines.extend([f"- {entry}" for entry in REMOVED_RECIPE_LINES])
    lines.append("")

    for recipe_name in recipe_names:
        recipe_id = normalize_key(recipe_name)
        guide = recipe_guides[recipe_id]
        lines.append(recipe_name)
        lines.append("")
        for index, step in enumerate(guide["steps"], start=1):
            lines.append(f"{index}. {step}")
        lines.append("")
        lines.append(f"Schwierigkeit {difficulty_dots(guide['difficulty'])}")
        lines.append(guide["tip"])
        lines.append("")
        lines.append("")

    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def build_summary_md(recipe_names: list[str]):
    original_images, optimized_images, joker_images = get_image_stats()
    path = OUT / "DOCUMENT_UPDATE_SUMMARY.md"
    lines = [
        "# Dokumentenupdate Giro: Pasta Night",
        "",
        f"- Stand: {TODAY.isoformat()}",
        "- Quelle: aktueller App-Stand im Projekt Giro: Pasta Night",
        "- Timer wurde in allen neuen Dokumenten weiter als entfernt / ausser Scope behandelt.",
        f"- Aktive Rezepte ohne Joker: {len(recipe_names)}",
        f"- Rezeptbilder: {original_images} Rezeptbilder + {joker_images} Joker-Bild",
        f"- Optimierte Web-Bilder: {optimized_images}",
        "",
        "## Neue Dateien",
        "- App_Architektur_v1.4.pdf",
        "- App_Konzept_v1.4.pdf",
        "- Architekturkarte v1.3.pdf",
        "- Aufgabenliste v1.4.xlsx",
        "- Mengenmodell Portionsbasis v1.4.xlsx",
        "- Pasta Saucen v2.3.txt",
        "- Regelblatt_v1.4.pdf",
        "- Rezeptliste v1.3.pdf",
        "- Zutatenliste v1.2.pdf",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def main():
    recipes_data = parse_recipes_data()
    recipe_guides = parse_recipe_guides()
    recipe_names = parse_pdf_recipe_names()
    validate_grouped_recipe_list(recipe_names)

    build_architecture_pdf(recipe_names)
    build_concept_pdf()
    build_architecture_card_pdf()
    build_rules_pdf()
    build_recipe_list_pdf(recipe_names)
    build_task_list_xlsx()
    build_amount_model_xlsx(recipes_data)
    build_ingredient_list_pdf(recipes_data)
    build_pasta_sauces_text(recipe_names, recipe_guides)
    build_summary_md(recipe_names)

    print(f"Generated files in {OUT}")
    for file in sorted(OUT.iterdir()):
        print(file.name)


if __name__ == "__main__":
    main()
